"""
Universal Report Loader

ML export files mix metadata (Client:, Program:, summary stats like
"Overview") with one or more data tables, and the header row's position
shifts depending on how much metadata precedes it. Rather than assuming a
fixed layout, this module scans the raw rows of any CSV/Excel export,
splits them into blocks on blank lines, and classifies each block as either
metadata (key/value pairs) or a data table (a header row followed by rows
of matching width).

The largest table block becomes the returned DataFrame. Actual parsing of
that table is delegated to pandas' native CSV/Excel readers (via skiprows/
nrows) so dtype inference stays identical to a hand-targeted read. All
metadata found along the way (Client, Program, summary counts, etc.) is
attached to the result as df.attrs["metadata"].
"""

import csv

import pandas as pd
from openpyxl import load_workbook


def _read_csv_rows(file):

    with open(file, "r", encoding="utf-8", errors="ignore") as f:
        return list(csv.reader(f))


def _read_excel_rows(file):

    wb = load_workbook(file, read_only=True, data_only=True)

    ws = wb.active

    rows = [
        ["" if cell is None else str(cell) for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]

    wb.close()

    return rows


def _is_blank(row):

    return all(str(cell).strip() == "" for cell in row)


def _as_kv_pair(cells):

    """
    Matches a metadata line such as ["Client: Deltek"] or
    ["Client", " Deltek"] and returns (key, value). Returns None
    if the cells don't look like a key/value pair.
    """

    if len(cells) == 1 and ":" in cells[0]:

        key, _, value = cells[0].partition(":")

        return key.strip(), value.strip()

    if len(cells) == 2:

        return cells[0].strip().rstrip(":"), cells[1].strip()

    return None


def _canonical_key(key):

    lowered = key.lower()

    if lowered.startswith("client"):
        return "Client"

    if lowered.startswith("program"):
        return "Program"

    return key


def _find_blocks(rows):

    """Groups consecutive non-blank rows into (start_index, rows) blocks."""

    blocks = []

    current = []
    start = 0

    for i, row in enumerate(rows):

        if _is_blank(row):

            if current:
                blocks.append((start, current))
                current = []

            continue

        if not current:
            start = i

        current.append(row)

    if current:
        blocks.append((start, current))

    return blocks


def _classify_block(start, block):

    """
    Returns ("metadata", {...}) for a block made entirely of key/value
    lines (and optional single-cell section titles like "Overview"),
    or ("table", header_index, data_row_count) for a block that contains
    a real header row followed by data.
    """

    kv = {}
    is_kv_block = True

    for row in block:

        cells = [c for c in row if str(c).strip() != ""]

        if not cells:
            continue

        pair = _as_kv_pair(cells)

        if pair:
            kv[_canonical_key(pair[0])] = pair[1]

        elif len(cells) == 1:
            continue  # section title, e.g. "Overview"

        else:
            is_kv_block = False
            break

    if is_kv_block and kv:
        return ("metadata", kv, None)

    for i, row in enumerate(block):

        cells = [c for c in row if str(c).strip() != ""]

        if len(cells) >= 2:
            return ("table", start + i, len(block) - i - 1)

    return (None, None, None)


def _parse(rows, file, reader):

    metadata = {}
    candidates = []

    for start, block in _find_blocks(rows):

        kind, a, b = _classify_block(start, block)

        if kind == "metadata":
            metadata.update(a)

        elif kind == "table" and b > 0:
            candidates.append((a, b))

    if not candidates:
        raise ValueError(f"Could not detect a data table in {file}")

    header_index, data_row_count = max(candidates, key=lambda c: c[1])

    df = reader(file, header_index, data_row_count)

    df.columns = df.columns.str.strip()

    df.attrs["metadata"] = metadata

    print(
        f"{file} -> header row {header_index}, "
        f"{len(df)} rows, {len(df.columns)} columns"
    )

    return df


def load_csv(file):

    rows = _read_csv_rows(file)

    return _parse(
        rows,
        file,
        reader=lambda f, skip, n: pd.read_csv(
            f, skiprows=skip, nrows=n, encoding="utf-8", low_memory=False
        )
    )


def load_excel(file):

    rows = _read_excel_rows(file)

    return _parse(
        rows,
        file,
        reader=lambda f, skip, n: pd.read_excel(f, skiprows=skip, nrows=n)
    )
